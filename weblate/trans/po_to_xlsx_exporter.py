# -*- coding: utf-8 -*-
#
# Copyright © 2012 - 2016 Michal Čihař <michal@cihar.com>
#
# This file is part of Weblate <https://weblate.org/>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
"""Exporter using polib and openpyxl"""

import os
import re

from polib import *

from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils import *

class PoToXlsxExporter(object):
    
    # It's a bit artificial to have all this as a class, but it's more convenient, 
    # until we solve the horrible issue that the below is largely copy/pasted from nimp localization...

    @staticmethod
    def export(po_file, repo_last_revision):
        po_file_base, po_file_ext = os.path.splitext(po_file)
        xlsx_file = po_file_base + '.xlsx'
        po_data = []
        po_data.append(pofile(po_file, wrapwidth=-1))
        wb = PoToXlsxExporter.build_workbook(po_data, repo_last_revision)
        wb.save(xlsx_file)
        return xlsx_file

    @staticmethod
    def export_multiple(po_files, repo_last_revision):
        for po_file in po_files:
            po_file_base, po_file_ext = os.path.splitext(po_file)
            xlsx_file = po_file_base + '.all.xlsx' # yeah I should remove lang in file name but this is only a temp file...
            break
        po_data = []
        for po_file in po_files:
            po_data.append(pofile(po_file, wrapwidth=-1))
        wb = PoToXlsxExporter.build_workbook(po_data, repo_last_revision)
        wb.save(xlsx_file)
        return xlsx_file

    @staticmethod
    def sort_by_lang(po_data):
        if len(po_data) > 1:
            # It's vital to have English first if present (then the rest can be sorted alphabetically)
            def _sort_by_lang(po_data_item):
                lang = PoToXlsxExporter.get_trans_column_title(po_data_item, True)
                if lang.startswith('en'): 
                    lang = '#' + lang # puts it first
                return lang
            po_data.sort(key = _sort_by_lang)

    @staticmethod
    def build_workbook(po_data, repo_last_revision=None):
        PoToXlsxExporter.sort_by_lang(po_data)
        wb, workbook_writing_instructions = PoToXlsxExporter.init_workbook(po_data, repo_last_revision)
        PoToXlsxExporter.fill_metadata(po_data, wb['metadata'], workbook_writing_instructions)
        i = 0
        for po_entry in po_data[0]: 
            if not po_entry.obsolete:
                PoToXlsxExporter.add_po_entry(po_entry, wb['data'], i, workbook_writing_instructions, PoToXlsxExporter.get_trans_column_title(po_data[0], len(po_data) > 1))
                i += 1
        if len(po_data) > 1:
            # now for all other languages, there is only one column to fill
            for i in range(1, len(po_data)):
                for po_entry in po_data[i]: 
                    if not po_entry.obsolete:
                        PoToXlsxExporter.fill_other_trans(po_entry, wb['data'], workbook_writing_instructions, PoToXlsxExporter.get_trans_column_title(po_data[i], len(po_data) > 1))
        if len(po_data) == 1: # (multiple po: just skip obsolete data for now)
            i = 0
            for po_entry in po_data[0].obsolete_entries(): 
                PoToXlsxExporter.add_po_entry(po_entry, wb['obsolete data'], i, workbook_writing_instructions, PoToXlsxExporter.get_trans_column_title(po_data[0], len(po_data) > 1))
                i += 1
        PoToXlsxExporter.finalize_workbook(wb)
        return wb

    @staticmethod
    def fill_metadata(po_data, ws, workbook_writing_instructions):
        # NB: when writing a multi-lang xlsx, only keep a few fields
        fields_multi_lang = ['Project-Id-Version', 'Report-Msgid-Bugs-To', 'POT-Creation-Date']
        i = workbook_writing_instructions[ws.title]['first_data_row']
        for name, value in po_data[0].ordered_metadata():
            if len(po_data) == 1 or name in fields_multi_lang:
                ws.cell(row=i, column=1).value = name
                ws.cell(row=i, column=2).value = value
                i += 1
        ws.cell(row=i, column=1).value = 'repo_last_revision'
        ws.cell(row=i, column=2).value = workbook_writing_instructions['repo_last_revision']

    @staticmethod
    def get_trans_column_titles(po_data):
        trans_column_titles = []
        for po_data_item in po_data:
            trans_column_titles.append(PoToXlsxExporter.get_trans_column_title(po_data_item, len(po_data) > 1))
        return trans_column_titles

    @staticmethod
    def get_trans_column_title(po_data, multiple_lang):
        if multiple_lang:
            for name, value in po_data.ordered_metadata():
                if name == 'Language':
                    return value
        else:
            return 'Translation'

    @staticmethod
    def init_workbook(po_data, repo_last_revision=None):
        wb = Workbook()
        workbook_writing_instructions = {}
        workbook_writing_instructions['repo_last_revision'] = repo_last_revision
        # The below will be helpful when exporting more than one language
        # Assumption though here: first language MUST be en and we have this weird situation where we
        # try to write a multi-lang. xlsx but actually en po and foreign po have a different set of (id, ctxt)
        # due to the en tweaks becoming "sources" for foreign pos... so it's a quite messy multi-lang. xlsx anyway
        workbook_writing_instructions['data_line_dictionary'] = {}
        trans_column_titles = PoToXlsxExporter.get_trans_column_titles(po_data)
        # data sheet
        dws = wb.active
        dws.title = 'data'
        do_write_data = False
        dne_column_titles_set = set()
        for po_entry in po_data[0]: 
            if not po_entry.obsolete:
                do_write_data = True
                PoToXlsxExporter.update_dne_columns(po_entry, dne_column_titles_set)
        if do_write_data:
            workbook_writing_instructions[dws.title] = {}
            workbook_writing_instructions[dws.title]['first_data_row'] = 2
            dne_column_titles = list(dne_column_titles_set)
            dne_column_titles.sort()
            workbook_writing_instructions[dws.title]['column_key'] = PoToXlsxExporter.write_header_row(dws, 1, dne_column_titles, trans_column_titles)
        # metadata sheet
        hws = wb.create_sheet()
        hws.title = 'metadata'
        workbook_writing_instructions[hws.title] = {}
        workbook_writing_instructions[hws.title]['first_data_row'] = 1
        # obsolete data sheet (when applicable) 
        # (multiple po: just skip obsolete data for now, as it's even more unreasonable to assume
        #  identical structure across languages for obsolete data than it is for regular data)
        if len(po_data) == 1:
            do_write_obsolete_data = False
            dne_column_titles_set = set()
            for po_entry in po_data[0].obsolete_entries():
                do_write_obsolete_data = True
                PoToXlsxExporter.update_dne_columns(po_entry, dne_column_titles_set)
            if do_write_obsolete_data:
                ows = wb.create_sheet()
                ows.title = 'obsolete data'
                workbook_writing_instructions[ows.title] = {}
                workbook_writing_instructions[ows.title]['first_data_row'] = 2
                dne_column_titles = list(dne_column_titles_set)
                dne_column_titles.sort()
                workbook_writing_instructions[ows.title]['column_key'] = PoToXlsxExporter.write_header_row(ows, 1, dne_column_titles, trans_column_titles)
        return wb, workbook_writing_instructions

    @staticmethod
    def write_header_row_part(ws, i, j, column_key, prefix_for_column_key, headers):
        header_fill = PatternFill(fill_type='solid', fgColor = 'FFFFFF4D' if prefix_for_column_key == '' else 'FFB3B3FF')
        # different color for Regular data (classic po) vs. Additional data (DNE data, embedded in comment)
        for header in headers:
            ws.cell(row=i, column=j).value = header
            ws.cell(row=i, column=j).fill = header_fill
            if not column_key is None:
                column_key[prefix_for_column_key + header] = j
            j += 1
        return j

    @staticmethod
    def write_header_row(ws, i, dne_column_titles, trans_column_titles):
        column_key = {}
        current_column = 1
        # Regular data (classic po): the important ones
        regular_important_columns = ['Source'] + trans_column_titles + ['Context']
        current_column = PoToXlsxExporter.write_header_row_part(ws, i, current_column, column_key, '', regular_important_columns)
        # Additional data (DNE data, embedded in comment)
        # (we consider that pretty important so show it before other classic po columns)
        current_column = PoToXlsxExporter.write_header_row_part(ws, i, current_column, column_key, '[DNE]', dne_column_titles)
        # Regular data (classic po): the less important ones
        regular_less_important_columns = ['Occurrences', 'Flags', 'Translator Comment', 
                                          'Previous Source', 'Previous Context', 
                                          'Comment']
        # Comment (raw) is not considered important because it bears data that is not relevant 
        # for translation/recording ('key' unique id), or data repeated in occurrences, or
        # additional data visible as unique columns, etc. So we ditch it at the end
        current_column = PoToXlsxExporter.write_header_row_part(ws, i, current_column, column_key, '', regular_less_important_columns)
        # ignore_for_now  = ['Source Plural', 'Translation Plural', 'Previous Source Plural', 'Line Number']
        return column_key

    @staticmethod
    def update_dne_columns(po_entry, dne_columns):
        dne_columns.update(set(PoToXlsxExporter.analyze_raw_comment(po_entry.comment).keys()))

    @staticmethod
    def analyze_raw_comment(comment):
        dict = {}
        # Empty additional data (trailing space seems to kinda disappear at some point in that case, so...):
        p = re.compile(r'^\[AdditionalData\] (?P<key>[^:]*):( ?)$', re.MULTILINE)
        for m in re.finditer(p, comment):
            dict[m.group('key')] = None
        # Additional data with a value:
        p = re.compile(r'^\[AdditionalData\] (?P<key>[^:]*): (?P<value>[^\n]+)$', re.MULTILINE)
        for m in re.finditer(p, comment):
            dict[m.group('key')] = m.group('value')
        return dict

    @staticmethod
    def finalize_worksheet(ws):
        if ws.title == 'metadata':
            PoToXlsxExporter.finalize_metadata_worksheet(ws)
        else:
            PoToXlsxExporter.finalize_data_worksheet(ws)

    @staticmethod
    def finalize_data_worksheet(ws):
        # TODO: either get rid of workbook_writing_instructions's 'first_data_row' or use it here...
        header_font = Font(bold=True)
        for j in range(1, ws.max_column + 1):
            ws.cell(row=1, column=j).font = header_font
        ws.auto_filter.ref = "%s:%s" % (get_column_letter(1), get_column_letter(ws.max_column))
        ws.freeze_panes = ws['A2']

    @staticmethod
    def finalize_metadata_worksheet(ws):
        header_font = Font(bold=True)
        for i in range(1, ws.max_row + 1):
            ws.cell(row=i, column=1).font = header_font

    @staticmethod
    def finalize_workbook(wb):
        for ws in wb.worksheets:
            PoToXlsxExporter.finalize_worksheet(ws)

    @staticmethod
    def inject_value(prefix_for_column_key, key, value, ws, i, column_key):
        ws.cell(row=i, column=column_key[prefix_for_column_key + key]).value = value

    @staticmethod
    def add_po_entry(po_entry, ws, i, workbook_writing_instructions, trans_column_title):
        def _format_flags(flags):
            return '\n'.join(flags)

        def _format_occurrence(fpath, lineno):
            if lineno:
                return '%s:%s' % (fpath, lineno)
            else:
                return fpath

        def _format_occurrences(occurrences):
            return '\n'.join(_format_occurrence(fpath, lineno) for fpath, lineno in occurrences)

        real_i = workbook_writing_instructions[ws.title]['first_data_row'] + i
        column_key = workbook_writing_instructions[ws.title]['column_key']
        PoToXlsxExporter.inject_value('', 'Source', po_entry.msgid, ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', trans_column_title, po_entry.msgstr, ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Context', po_entry.msgctxt, ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Comment', po_entry.comment, ws, real_i, column_key)
        # Fairly twisted rules below because of the multi-lang. & tweaks mess:
        tweaked_msgid = po_entry.msgid if po_entry.msgstr == '' else po_entry.msgstr
        if (po_entry.msgctxt, tweaked_msgid) not in workbook_writing_instructions['data_line_dictionary']:
            workbook_writing_instructions['data_line_dictionary'][(po_entry.msgctxt, tweaked_msgid)] = set()
        workbook_writing_instructions['data_line_dictionary'][(po_entry.msgctxt, tweaked_msgid)].update(set([real_i]))
        PoToXlsxExporter.inject_value('', 'Translator Comment', po_entry.tcomment, ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Occurrences', _format_occurrences(po_entry.occurrences), ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Flags', _format_flags(po_entry.flags), ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Previous Source', po_entry.previous_msgid, ws, real_i, column_key)
        PoToXlsxExporter.inject_value('', 'Previous Context', po_entry.previous_msgctxt, ws, real_i, column_key)
        # Ignore for now
        # PoToXlsxExporter.inject_value('', 'Source Plural', po_entry.msgid_plural, ws, real_i, column_key)
        # PoToXlsxExporter.inject_value('', 'Translation Plural', po_entry.msgstr_plural, ws, real_i, column_key)
        # PoToXlsxExporter.inject_value('', 'Previous Source Plural', po_entry.previous_msgid_plural, ws, real_i, column_key)
        # PoToXlsxExporter.inject_value('', 'Line Number', po_entry.linenum, ws, real_i, column_key)
        for key, value in PoToXlsxExporter.analyze_raw_comment(po_entry.comment).items():
            PoToXlsxExporter.inject_value('[DNE]', key, value, ws, real_i, column_key)

    @staticmethod
    def fill_other_trans(po_entry, ws, workbook_writing_instructions, trans_column_title):
        column_key = workbook_writing_instructions[ws.title]['column_key']
        for real_i in workbook_writing_instructions['data_line_dictionary'][(po_entry.msgctxt, po_entry.msgid)]:
            PoToXlsxExporter.inject_value('', trans_column_title, po_entry.msgstr, ws, real_i, column_key)

    @staticmethod
    def parse_workbook(wb, simple_mode, alt_translation_column_name=None):
        po_data, workbook_reading_instructions = PoToXlsxExporter.init_po(wb, alt_translation_column_name)
        repo_old_revision = PoToXlsxExporter.read_metadata(po_data, wb, workbook_reading_instructions)
        PoToXlsxExporter.read_data(po_data, wb, workbook_reading_instructions, simple_mode)
        if not simple_mode:
            PoToXlsxExporter.read_obsolete_data(po_data, wb, workbook_reading_instructions)
        return po_data, repo_old_revision

    @staticmethod
    def init_po_using_sheet(wb, sheetname, header_row, first_data_row, workbook_reading_instructions, alt_translation_column_name=None):
        if sheetname in wb.sheetnames:
            workbook_reading_instructions[sheetname] = {}
            workbook_reading_instructions[sheetname]['first_data_row'] = first_data_row
            workbook_reading_instructions[sheetname]['column_key'] = PoToXlsxExporter.read_header_row(wb[sheetname], header_row)
            if (alt_translation_column_name is not None 
                and 'Translation' not in workbook_reading_instructions[sheetname]['column_key'] 
                and alt_translation_column_name in workbook_reading_instructions[sheetname]['column_key']):
                workbook_reading_instructions[sheetname]['column_key']['Translation'] = workbook_reading_instructions[sheetname]['column_key'][alt_translation_column_name]

    @staticmethod
    def init_po(wb, alt_translation_column_name=None):
        po_data = POFile(wrapwidth=-1)
        workbook_reading_instructions = {}
        # data sheet
        PoToXlsxExporter.init_po_using_sheet(wb, 'data', 1, 2, workbook_reading_instructions, alt_translation_column_name)
        # metadata sheet
        if 'metadata' in wb.sheetnames:
            workbook_reading_instructions['metadata'] = {}
            workbook_reading_instructions['metadata']['first_data_row'] = 1
        # obsolete data sheet (when applicable)
        PoToXlsxExporter.init_po_using_sheet(wb, 'obsolete data', 1, 2, workbook_reading_instructions, alt_translation_column_name)
        return po_data, workbook_reading_instructions

    @staticmethod
    def read_header_row(ws, i):
        column_key = {}
        current_column = 1
        for j in range(1, ws.max_column + 1):
            column_key[ws.cell(row=i, column=j).value] = j
        return column_key

    @staticmethod
    def read_metadata(po_data, wb, workbook_reading_instructions):
        po_data.metadata = {}
        if 'metadata' in wb.sheetnames:
            ws = wb['metadata']
            for i in range(workbook_reading_instructions[ws.title]['first_data_row'], ws.max_row + 1):
                name = ws.cell(row=i, column=1).value
                value = ws.cell(row=i, column=2).value
                if name == 'repo_last_revision':
                    repo_old_revision = value
                else:
                    po_data.metadata[name] = value
            return repo_old_revision

    @staticmethod
    def read_any_data(po_data, ws, workbook_reading_instructions, obsolete, simple_mode=False):
        for i in range(workbook_reading_instructions[ws.title]['first_data_row'], ws.max_row + 1):
            po_entry = POEntry()
            column_key = workbook_reading_instructions[ws.title]['column_key']
            po_entry.msgid = PoToXlsxExporter.read_value(ws, i, 'Source', column_key)
            po_entry.msgstr = PoToXlsxExporter.read_value(ws, i, 'Translation', column_key)
            po_entry.msgctxt = PoToXlsxExporter.read_value(ws, i, 'Context', column_key, keep_none=True)
            po_entry.obsolete = obsolete
            po_entry.tcomment = PoToXlsxExporter.read_value(ws, i, 'Translator Comment', column_key)
            flags = PoToXlsxExporter.read_value(ws, i, 'Flags', column_key)
            if flags:
                po_entry.flags = flags.split('\n')
            po_entry.previous_msgid = PoToXlsxExporter.read_value(ws, i, 'Previous Source', column_key)
            po_entry.previous_msgctxt = PoToXlsxExporter.read_value(ws, i, 'Previous Context', column_key)
            if not simple_mode:
                # Typically (upload scenario in Weblate),
                # (extracted) comments and occurrences are not at all needed in the output file
                # if said po file is just used for a merge towards a more regular file
                comment = PoToXlsxExporter.read_value(ws, i, 'Comment', column_key)
                po_entry.comment = comment
                occurrences = PoToXlsxExporter.read_value(ws, i, 'Occurrences', column_key)
                if occurrences:
                    occurrences_list = occurrences.split('\n')
                    for occurrence in occurrences_list:
                        fpath = occurrence
                        lineno = ''
                        occurrence_fpath_lineno = occurrence.split(':')
                        if len(occurrence_fpath_lineno) == 2 and occurrence_fpath_lineno[1].isdigit():
                            fpath = occurrence_fpath_lineno[0]
                            lineno = occurrence_fpath_lineno[1]
                        po_entry.occurrences.append((fpath, lineno)) 
                        # NB: some issues remain with colon (:) character in occurrences, but they don't seem critical
                        # Maybe this colon (:) character in occurrences should be handled differently as early as
                        # in ManifestToPot so as to not disturb polib
            # Ignore for now
            # po_entry.msgid_plural = PoToXlsxExporter.read_value(ws, i, 'Source Plural', column_key)
            # po_entry.msgstr_plural = PoToXlsxExporter.read_value(ws, i, 'Translation Plural', column_key)
            # po_entry.previous_msgid_plural = PoToXlsxExporter.read_value(ws, i, 'Previous Source Plural', column_key)
            # po_entry.linenum = PoToXlsxExporter.read_value(ws, i, 'Line Number', column_key)
            po_data.append(po_entry)

    @staticmethod
    def read_value(ws, i, key, column_key, keep_none=False):
        value = ws.cell(row=i, column=column_key[key]).value
        if keep_none:
            return value
        else:
            return '' if value is None else value

    @staticmethod
    def read_data(po_data, wb, workbook_reading_instructions, simple_mode):
        if 'data' in wb.sheetnames:
            PoToXlsxExporter.read_any_data(po_data, wb['data'], workbook_reading_instructions, False, simple_mode)

    @staticmethod
    def read_obsolete_data(po_data, wb, workbook_reading_instructions):
        if 'obsolete data' in wb.sheetnames:
            PoToXlsxExporter.read_any_data(po_data, wb['obsolete data'], workbook_reading_instructions, True)

def xlsx_to_po(xlsx_file, alt_translation_column_name=None):
    xlsx_file_base, xlsx_file_ext = os.path.splitext(xlsx_file)
    po_file = xlsx_file_base + '.po'
    wb = load_workbook(xlsx_file)
    po_data, repo_old_revision = PoToXlsxExporter.parse_workbook(wb, True, alt_translation_column_name)
    po_data.save(fpath=po_file)
    return po_file, repo_old_revision
